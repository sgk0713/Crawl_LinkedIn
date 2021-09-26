# encoding=utf8

# python 3.6

import csv
import os
import random
import time

import openpyxl
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys

import config_constants
import logger

mergedDir = "./merged"
blankEmailDir = "./blankEmail"
errorDir = "./error"
crawledDir = "./crawled"

previousCompletedFileName = f'test.xlsx'  # replace your last completed crawled file name. if not exist, just put any name you want(not blank)
global currentTime
global mergedFileName
global connectionFileName
global blankEmailFileName
global crawledFileName
global errorFileName
global mergeErrorFileName

Login_Button_XPATH = '//*[@id="main-content"]/section[1]/div[2]/form/button'
SEARCH_FIELD_XPATH = '//*[@id="mn-connections-search-input"]'
SEARCH_RESULT_FIRST_ITEM_SELECTOR = 'span.mn-connection-card__name.t-16.t-black.t-bold'

NAME_FIELD_SELECTOR = 'div.display-flex.justify-space-between.pt2 > div > div:nth-child(1) > h1'
CONTACT_FIELD_SELECTOR = 'div.pb2 > span.pv-text-details__separator.t-black--light > a'

WEBSITE_ROOT_IN_CONTACT_SELECTOR = 'section.pv-contact-info__contact-type.ci-websites > ul li'
EMAIL_FIELD_IN_CONTACT_SELECTOR = 'div > section.pv-contact-info__contact-type.ci-email > div > a'

names = []  # it will be filled from connections file
namesRowMap = {}  # key: name, value: rowIndex

CONST_ID = config_constants.CONST_ID  # read from config_constants.py. put your infos at config_constants.py file
CONST_PW = config_constants.CONST_PW
DRIVER_PATH = config_constants.CONST_DRIVER_PATH

global driver


def initChromeDriver():
    global driver
    driver = webdriver.Chrome(DRIVER_PATH)
    driver.implicitly_wait(1)


def login():
    if driver is None:
        logger.log("driver is NONE")
        exit(1)
    driver.get('https://www.linkedin.com/')
    driver.find_element_by_name('session_key').send_keys(CONST_ID)
    driver.find_element_by_name('session_password').send_keys(CONST_PW)
    driver.find_element_by_xpath(Login_Button_XPATH).click()


def moveToConnectionPage():
    if driver is None:
        logger.log("driver is NONE")
        exit(1)
    driver.get('https://www.linkedin.com/mynetwork/invite-connect/connections/')


def startCrawling(start: int = 0, itemCount: int = 100):
    max = len(names)
    start = start
    end = min(start + itemCount, max)

    isFileExist = os.path.isfile(crawledFileName)

    f = open(crawledFileName, 'a', encoding='utf-8', newline='')
    wr = csv.writer(f)

    errorFile = open(errorFileName, 'a', encoding='utf-8', newline='')
    ewr = csv.writer(errorFile)

    if not isFileExist:
        wr.writerow(['Number', 'TypedName', 'ProfileName', 'Email', 'HomePage'])
        ewr.writerow(['i'])

    logger.log(f'total : {max} start : {start} end : {end}')

    initChromeDriver()
    login()
    time.sleep(10)
    moveToConnectionPage()
    time.sleep(10)

    breakTimes = [1, 1, 1]

    for i in range(start, end):
        time.sleep(breakTimes[i % 3])
        if i % 3 == 0:
            breakTimes = getBreakTimes()
        try:
            search_name(names[i])
            time.sleep(3)
            clickFirstItemAndMoveToPage()
            time.sleep(3)

            name = driver.find_element_by_css_selector(NAME_FIELD_SELECTOR).text
            clickContactField()
        except Exception as e:
            logger.log(f'search {i} {names[i]}')
            logger.log(e)
            driver.switch_to.window(driver.window_handles[0])
            ewr.writerow(['i == %d' % i, names[i]])
            break
        homepages = []
        email = ""
        try:
            lists = driver.find_elements_by_css_selector(WEBSITE_ROOT_IN_CONTACT_SELECTOR)
            for list in lists:
                homepages.append(list.find_element_by_css_selector('div a').get_attribute("href"))
        except Exception as e:
            logger.log("NO Website Field Exception")
            logger.log(e)
            homepages = None
        try:
            email = driver.find_element_by_css_selector(EMAIL_FIELD_IN_CONTACT_SELECTOR).text
        except Exception as e:
            logger.log('No Email')
            logger.log(e)
            email = '-'
        try:
            time.sleep(1)
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            logger.log(f'#{i} {name} >> {email}')
            list = []
            list.append(i)  # 'Number'
            list.append(names[i])  # 'TypedName'
            list.append(name)  # 'ProfileName'
            list.append(email)  # 'Email'
            if homepages is not None:  # 'HomePage'
                for homepage in homepages:
                    list.append(homepage)
            wr.writerow(list)
        except Exception as e:
            logger.log('homepages error')
            logger.log(e)
            driver.switch_to.window(driver.window_handles[0])
            ewr.writerow(['i == %d' % i])
            continue
    f.close()
    errorFile.close()


def clickContactField():
    contactField = driver.find_element_by_css_selector(
        CONTACT_FIELD_SELECTOR)
    ActionChains(driver).click(contactField).perform()


def clickFirstItemAndMoveToPage():
    element = driver.find_element_by_css_selector(SEARCH_RESULT_FIRST_ITEM_SELECTOR)
    ActionChains(driver).key_down(Keys.COMMAND).click(element).key_up(Keys.COMMAND).perform()
    driver.switch_to.window(driver.window_handles[1])


def search_name(name: str):
    driver.find_element_by_xpath(SEARCH_FIELD_XPATH).clear()
    driver.find_element_by_xpath(SEARCH_FIELD_XPATH).clear()
    driver.find_element_by_xpath(SEARCH_FIELD_XPATH).send_keys(name)


def isBlankOrNone(myString):
    if type(myString) is str and myString.strip():
        # myString is not None AND myString is not empty or blank
        return False
    # myString is None OR myString is empty or blank
    return True


def mergeConnectionsFile():
    if not os.path.isfile(connectionFileName):
        logger.log("no connection file")
        exit(1)
    else:
        logger.log("start read")

    isPreviousFileExist = os.path.isfile(previousCompletedFileName)

    if not isPreviousFileExist:
        wt = Workbook()
        wtactive = wt.active
        wtactive.title = 'Sheet1'
        titles = ["First Name",
                  "Last Name",
                  "Full Name",
                  "Email Address",
                  "Company",
                  "Position",
                  "Connected On",
                  "HomePage"]
        for (index, title) in enumerate(titles):
            wtactive.cell(column=index + 2, row=2).value = title
        wt.save(previousCompletedFileName)
        wt.close()

    wb = openpyxl.load_workbook(previousCompletedFileName)
    wbactive = wb.active
    firstLastPositionList = set()
    for line in wbactive.iter_rows(values_only=True):
        try:
            if isBlankOrNone(line[5]):
                company = ''
            else:
                company = line[5]
            name = line[1] + line[2] + company
            firstLastPositionList.add(name)  # first last companyname
        except Exception as e:
            logger.log("exception from reading previous file")
            logger.log(e)

    nameFile = open(connectionFileName, 'r', encoding='utf-8', newline='')
    connectionListReader = csv.reader(nameFile)
    startRead = False
    addedCount = 0
    for line in connectionListReader:
        try:
            if not line[0].strip() and not line[1].strip():
                continue
            if startRead:
                name = line[0] + line[1] + line[3]
                element = firstLastPositionList.intersection([name])
                needCopy = len(element) == 0
                if needCopy:
                    row = ['',
                           line[0],
                           line[1],
                           line[0] + ' ' + line[1]
                           ]

                    for index, item in enumerate(line):
                        if index > 1:
                            row.append(item)
                    wbactive.append(row)
                    addedCount += 1

            if line[0] == 'First Name':
                startRead = True
        except Exception as e:
            logger.log("exception from merging connections and previous files")
            logger.log(e)

    wb.save(mergedFileName)
    wb.close()
    nameFile.close()
    logger.log(f'addedCount from connection: {addedCount}')


def extractCrawlingList():
    if not os.path.isfile(mergedFileName):
        logger.log("initializeUpdateList no file")
        exit(1)
    else:
        logger.log("initializeUpdateList start read")

    wb = openpyxl.load_workbook(mergedFileName)
    wbactive = wb.active

    blankEmailFile = open(blankEmailFileName, 'w', encoding='utf-8', newline='')
    writer = csv.writer(blankEmailFile)

    blankEmailCount = 0
    for row in range(3, wbactive.max_row + 1):  # cells starts index 1, not 0
        try:
            email = wbactive.cell(column=5, row=row).value
            fullName = wbactive.cell(column=4, row=row).value
            if isBlankOrNone(email) and not isBlankOrNone(fullName):
                blankEmailCount += 1
                names.append(fullName)
                writer.writerow([fullName])
                namesRowMap[fullName] = row
        except Exception as e:
            logger.log('exception from initializeUpdateList')
            logger.log(e)
    wb.save(mergedFileName)
    wb.close()
    blankEmailFile.close()

    logger.log(f'Blank Email Count: {blankEmailCount}')


def getBreakTimes():
    breakTimes = []
    firstBreakTime = random.randint(10, 20)
    secondBreakTime = random.randint(10, 45 - firstBreakTime)
    thirdBreakTime = 45 - secondBreakTime - firstBreakTime
    breakTimes.clear()
    breakTimes.append(firstBreakTime)
    breakTimes.append(secondBreakTime)
    breakTimes.append(thirdBreakTime)
    logger.log(f'breaktime reset {breakTimes[0]}, {breakTimes[1]}, {breakTimes[2]}')
    return breakTimes


def initDir():
    dirList = [mergedDir, blankEmailDir, errorDir, crawledDir]
    try:
        for dir in dirList:
            if not os.path.exists(dir):
                os.makedirs(dir)
    except OSError:
        logger.log('Error: Creating directory. ' + dir)


def updateMergedFile():
    if not os.path.isfile(crawledFileName) or not os.path.isfile(mergedFileName):
        logger.log("no file")
        exit(1)
    else:
        logger.log("start read")

    wb = openpyxl.load_workbook(mergedFileName)
    wbactive = wb.active
    crawledFile = open(crawledFileName, 'r', encoding='utf-8', newline='')
    mergeErrorFile = open(mergeErrorFileName, 'a', encoding='utf-8', newline='')

    wtr = csv.writer(mergeErrorFile)
    rdr = csv.reader(crawledFile)
    startRead = False
    for line in rdr:
        try:
            if not line[0].strip() and not line[1].strip():
                continue
            if startRead:
                # fullName = line[2]
                email = line[3]
                fullName = line[1]
                # email = line[2]
                homePages = []
                for index in range(4, len(line)):
                    homePages.append(line[index])
                if namesRowMap[fullName] is not None:
                    row = namesRowMap[fullName]
                    if not isBlankOrNone(wbactive.cell(column=5, row=row).value):
                        continue
                    wbactive.cell(column=5, row=row).value = email
                    for index in range(9, 9 + len(homePages)):
                        wbactive.cell(column=index, row=row).value = homePages[9 - index]
            if line[0] == 'Number':
                startRead = True
        except Exception as e:
            logger.log('updateMergedFile')
            wtr.writerow(line)
            logger.log(e)

    wb.save(mergedFileName)
    wb.close()
    crawledFile.close()
    mergeErrorFile.close()


def initFileNamesWithCurrentTime():
    global currentTime
    global mergedFileName
    global connectionFileName
    global blankEmailFileName
    global crawledFileName
    global errorFileName
    global mergeErrorFileName

    currentTime = time.strftime('%y%m%d_%I%M%S', time.localtime(time.time()))
    mergedFileName = f"{mergedDir}/Merged_{currentTime}.xlsx"
    connectionFileName = 'Connections.csv'
    blankEmailFileName = f'{blankEmailDir}/blank_email_{currentTime}.csv'
    crawledFileName = f'{crawledDir}/crawled_{currentTime}.csv'
    errorFileName = f'{errorDir}/error_{currentTime}.csv'
    mergeErrorFileName = f'{errorDir}/mergeError_{currentTime}.csv'


if __name__ == '__main__':
    initDir()

    initFileNamesWithCurrentTime()
    mergeConnectionsFile()
    extractCrawlingList()
    startCrawling(0, 100)
    updateMergedFile()
